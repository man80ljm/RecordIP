from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


class ExcelServiceError(Exception):
    """Excel 读写相关异常。"""


class ExcelService:
    """Excel 服务，负责历史记录存储与读取。"""

    HEADERS = [
        "记录时间",
        "主源IP",
        "网页源IP",
        "国家",
        "地区",
        "城市",
        "ISP",
        "组织",
        "AS",
        "时区",
        "双源一致",
        "是否变化",
        "备注",
    ]

    def ensure_workbook(self, excel_path: Path) -> None:
        """确保 Excel 文件存在且包含正确表头。"""
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        if not excel_path.exists():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "IP记录"
            sheet.append(self.HEADERS)
            try:
                workbook.save(excel_path)
            except PermissionError as exc:
                raise ExcelServiceError("无法创建 Excel 文件，文件可能被占用") from exc
            except OSError as exc:
                raise ExcelServiceError(f"创建 Excel 文件失败: {exc}") from exc
            return

        # 文件已存在时，检查表头并自动修复。
        try:
            workbook = load_workbook(excel_path)
            sheet = workbook.active
        except PermissionError as exc:
            raise ExcelServiceError("无法打开 Excel 文件，文件可能被占用") from exc
        except OSError as exc:
            raise ExcelServiceError(f"打开 Excel 文件失败: {exc}") from exc

        if sheet.max_row == 0:
            sheet.append(self.HEADERS)
        else:
            current_headers = [sheet.cell(row=1, column=i + 1).value for i in range(len(self.HEADERS))]
            if current_headers != self.HEADERS:
                for i, header in enumerate(self.HEADERS, start=1):
                    sheet.cell(row=1, column=i, value=header)

        try:
            workbook.save(excel_path)
        except PermissionError as exc:
            raise ExcelServiceError("保存 Excel 失败，文件可能被占用") from exc
        except OSError as exc:
            raise ExcelServiceError(f"保存 Excel 失败: {exc}") from exc

    def get_last_ips(self, excel_path: Path) -> dict[str, str]:
        """读取最后一条记录中的主源 IP 和网页源 IP。"""
        self.ensure_workbook(excel_path)

        try:
            workbook = load_workbook(excel_path)
            sheet = workbook.active
        except PermissionError as exc:
            raise ExcelServiceError("读取 Excel 失败，文件可能被占用") from exc
        except OSError as exc:
            raise ExcelServiceError(f"读取 Excel 失败: {exc}") from exc

        if sheet.max_row <= 1:
            return {"primary_ip": "", "web_ip": ""}

        # 从底部向上找最后一条有效记录（主源IP列2，网页源IP列3）。
        for row in range(sheet.max_row, 1, -1):
            primary = sheet.cell(row=row, column=2).value
            web = sheet.cell(row=row, column=3).value
            if primary or web:
                return {
                    "primary_ip": str(primary).strip() if primary else "",
                    "web_ip": str(web).strip() if web else "",
                }

        return {"primary_ip": "", "web_ip": ""}

    def get_last_ip(self, excel_path: Path) -> str:
        """向后兼容：返回最后记录的网页源 IP。"""
        return self.get_last_ips(excel_path).get("web_ip", "")

    def append_record(self, excel_path: Path, record: dict[str, Any]) -> None:
        """追加一条检测记录。"""
        self.ensure_workbook(excel_path)

        try:
            workbook = load_workbook(excel_path)
            sheet = workbook.active
        except PermissionError as exc:
            raise ExcelServiceError("写入 Excel 失败，文件可能被占用") from exc
        except OSError as exc:
            raise ExcelServiceError(f"写入 Excel 失败: {exc}") from exc

        sheet.append(
            [
                record.get("记录时间", ""),
                record.get("主源IP", ""),
                record.get("网页源IP", ""),
                record.get("国家", ""),
                record.get("地区", ""),
                record.get("城市", ""),
                record.get("ISP", ""),
                record.get("组织", ""),
                record.get("AS", ""),
                record.get("时区", ""),
                record.get("双源一致", ""),
                record.get("是否变化", ""),
                record.get("备注", ""),
            ]
        )

        try:
            workbook.save(excel_path)
        except PermissionError as exc:
            raise ExcelServiceError("保存 Excel 失败，文件可能被占用") from exc
        except OSError as exc:
            raise ExcelServiceError(f"保存 Excel 失败: {exc}") from exc
